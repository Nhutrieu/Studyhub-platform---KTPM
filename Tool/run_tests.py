import argparse
import os
import json
import shutil
import subprocess
from datetime import datetime, timezone
import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Tải cấu hình từ file .env
load_dotenv()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

if not os.path.exists(os.path.join(BASE_DIR, ".env")):
    print("\n⚠️  CẢNH BÁO: Không tìm thấy tệp cấu hình .env trong thư mục Tool/.")
    print("👉 Hãy tạo tệp .env (bằng cách sao chép từ .env.example) và điền cấu hình cá nhân của bạn.\n")

BASE_URL = os.getenv("BASE_URL", "http://localhost:8000")
PROJECT_NAME = os.getenv("PROJECT_NAME", "StudyHub")
JIRA_BASE_URL = os.getenv("JIRA_BASE_URL")
JIRA_EMAIL = os.getenv("JIRA_EMAIL")
JIRA_API_TOKEN = os.getenv("JIRA_API_TOKEN")
JIRA_PROJECT_KEY = os.getenv("JIRA_PROJECT_KEY", "SH")
JIRA_ISSUE_TYPE = os.getenv("JIRA_ISSUE_TYPE", "Bug")
COLLECTION_PATH = os.getenv("POSTMAN_COLLECTION_PATH", "postman/collections")

if os.path.isabs(COLLECTION_PATH):
    COLLECTION = COLLECTION_PATH
else:
    COLLECTION = os.path.join(BASE_DIR, COLLECTION_PATH)

REPORT_DIR = os.path.join(BASE_DIR, "postman", "reports", "local-run")
REPORT_JSON = os.path.join(REPORT_DIR, "report.json")
REPORT_MD = os.path.join(REPORT_DIR, "result.md")
REPORT_XLSX = os.path.join(REPORT_DIR, "report.xlsx")

os.makedirs(REPORT_DIR, exist_ok=True)

def run_newman_for_collections(collections, base_url):
    newman = shutil.which("newman")
    # Nếu không thấy global newman, thử tìm trong node_modules cục bộ của thư mục Tool
    if not newman:
        local_newman_bin = os.path.join(BASE_DIR, "node_modules", ".bin", "newman")
        # Trên Windows, nhị phân có thể là newman.cmd
        if os.name == "nt":
            local_newman_bin += ".cmd"
        if os.path.exists(local_newman_bin):
            newman = local_newman_bin

    if not newman:
        print("Không tìm thấy Newman. Vui lòng chạy 'npm install' hoặc 'npm install -g newman'")
        return False

    consolidated_data = {"run": {"executions": []}}
    temp_dir = os.path.join(REPORT_DIR, "temp")
    os.makedirs(temp_dir, exist_ok=True)
    success = False

    for idx, coll in enumerate(collections):
        temp_report_json = os.path.join(temp_dir, f"report_{idx}.json")
        cmd = [
            newman, "run", coll,
            "--env-var", f"baseUrl={base_url}",
            "--reporters", "cli,json",
            "--reporter-json-export", temp_report_json
        ]
        print(f"Running Newman for: {os.path.basename(coll)} ...")
        subprocess.run(cmd)
        if os.path.exists(temp_report_json):
            success = True
            try:
                with open(temp_report_json, "r", encoding="utf-8") as f:
                    rep_data = json.load(f)
                    executions = rep_data.get("run", {}).get("executions", [])
                    consolidated_data["run"]["executions"].extend(executions)
            except Exception as e:
                print(f"Lỗi khi đọc file báo cáo tạm: {e}")

    try:
        shutil.rmtree(temp_dir, ignore_errors=True)
    except Exception:
        pass

    if success:
        with open(REPORT_JSON, "w", encoding="utf-8") as f:
            json.dump(consolidated_data, f, indent=2, ensure_ascii=False)
        return True
    return False

def parse_report():
    if not os.path.exists(REPORT_JSON):
        return [], []
    with open(REPORT_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)
    results = []
    failures = []
    for ex in data.get("run", {}).get("executions", []):
        name = ex.get("item", {}).get("name", "Unknown Test")
        request = ex.get("request", {})
        method = request.get("method", "UNKNOWN")
        url_obj = request.get("url", {})
        url = url_obj.get("raw", "") if isinstance(url_obj, dict) else str(url_obj)
        response = ex.get("response", {})
        status = response.get("code", "NO RESPONSE")
        assertion_errors = []
        for assertion in ex.get("assertions", []):
            error = assertion.get("error")
            if error:
                assertion_errors.append(f"{assertion.get('assertion')}: {error.get('message')}")
        result = "FAIL" if assertion_errors or status not in [200, 201, 204] else "PASS"
        row = {
            "name": name,
            "method": method,
            "url": url,
            "status": status,
            "result": result,
            "message": "; ".join(assertion_errors)
        }
        results.append(row)
        if result == "FAIL":
            failures.append(row)
    return results, failures

def write_markdown(results, failures, created_issues):
    now = datetime.now(timezone.utc).isoformat()
    lines = [
        f"# {PROJECT_NAME} API Testing Report", "",
        f"Generated: {now}", "",
        "## Test Summary", "",
        "| Test Case | Method | URL | Status | Result | Message |",
        "|---|---|---|---|---|---|"
    ]
    for r in results:
        lines.append(f"| {r['name']} | {r['method']} | {r['url']} | {r['status']} | {r['result']} | {r['message']} |")
    lines.extend(["", "## Jira Issues", ""])
    if created_issues:
        for issue in created_issues:
            lines.append(f"- {issue}")
    else:
        lines.append("No Jira issues created.")
    with open(REPORT_MD, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

def write_excel(results, failures, created_issues):
    wb = Workbook()
    ws = wb.active
    ws.title = "API Summary"
    blue_fill = PatternFill("solid", fgColor="00008B")
    white_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    headers = ["No", "Function Name", "Passed", "Failed", "Untested", "N", "A", "B", "Total Test Cases"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border
    for index, r in enumerate(results, start=1):
        passed = 1 if r["result"] == "PASS" else 0
        failed = 1 if r["result"] == "FAIL" else 0
        ws.append([index, r["name"], passed, failed, 0, 0, 0, 0, 1])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = center
    widths = {"A": 8, "B": 35, "C": 12, "D": 12, "E": 12, "F": 8, "G": 8, "H": 8, "I": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    total_row = len(results) + 2
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = Font(bold=True)
    ws[f"C{total_row}"] = f"=SUM(C2:C{len(results)+1})"
    ws[f"D{total_row}"] = f"=SUM(D2:D{len(results)+1})"
    ws[f"E{total_row}"] = f"=SUM(E2:E{len(results)+1})"
    ws[f"I{total_row}"] = f"=SUM(I2:I{len(results)+1})"
    for cell in ws[total_row]:
        cell.border = border
        cell.alignment = center
    wb.save(REPORT_XLSX)

def create_jira_issue(failure):
    if not all([JIRA_BASE_URL, JIRA_EMAIL, JIRA_API_TOKEN, JIRA_PROJECT_KEY]):
        return None
    url = f"{JIRA_BASE_URL.rstrip('/')}/rest/api/3/issue"
    description_text = (
        f"Mô tả lỗi\nPhát hiện lỗi tự động khi chạy kiểm thử API [{failure['name']}].\n\n"
        f"Chi tiết lỗi:\n- Method: {failure['method']}\n- URL: {failure['url']}\n- Status: {failure['status']}\n- Lỗi: {failure['message']}"
    )
    payload = {
        "fields": {
            "project": {"key": JIRA_PROJECT_KEY},
            "summary": f"[{PROJECT_NAME.upper()} API TEST FAILED] {failure['name']}",
            "description": {
                "type": "doc",
                "version": 1,
                "content": [{"type": "paragraph", "content": [{"type": "text", "text": description_text}]}]
            },
            "issuetype": {"name": JIRA_ISSUE_TYPE}
        }
    }
    try:
        response = requests.post(
            url,
            json=payload,
            auth=(JIRA_EMAIL, JIRA_API_TOKEN),
            headers={"Accept": "application/json", "Content-Type": "application/json"}
        )
        if response.status_code in [200, 201]:
            return response.json().get("key")
        else:
            print(f"Lỗi khi log Bug Jira: HTTP {response.status_code} - {response.text}")
    except Exception as e:
        print(f"Lỗi kết nối tới Jira: {e}")
    return None

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--collection", default=COLLECTION, help="Path to Postman collection JSON or folder")
    parser.add_argument("--base-url", default=BASE_URL, help="Base URL of the API under test")
    args = parser.parse_args()

    col_path = args.collection if os.path.isabs(args.collection) else os.path.normpath(os.path.join(BASE_DIR, args.collection))
    collections = []
    if os.path.isdir(col_path):
        for root, dirs, files in os.walk(col_path):
            for file in files:
                if file.endswith(".json") and "environment" not in file.lower() and "global" not in file.lower():
                    collections.append(os.path.join(root, file))
    elif os.path.exists(col_path):
        collections.append(col_path)

    if not collections:
        print(f"Không tìm thấy file kịch bản test tại: {col_path}")
        return

    print(f"Đã tìm thấy {len(collections)} kịch bản kiểm thử.")
    if not run_newman_for_collections(collections, args.base_url):
        return

    results, failures = parse_report()
    created_issues = []
    if failures:
        print(f"Phát hiện {len(failures)} testcase lỗi.")
        for failure in failures:
            issue_key = create_jira_issue(failure)
            if issue_key:
                created_issues.append(issue_key)
                print(f"-> Đã tạo Jira Issue: {issue_key}")

    write_markdown(results, failures, created_issues)
    write_excel(results, failures, created_issues)
    
    print("\n========================================================================")
    print("Quá trình kiểm thử và tạo báo cáo hoàn tất!")
    print(f"- File JSON tổng hợp: {REPORT_JSON}")
    print(f"- File Báo cáo Excel: {REPORT_XLSX}")
    print(f"- File Báo cáo Markdown: {REPORT_MD}")
    print("========================================================================")

if __name__ == "__main__":
    main()
