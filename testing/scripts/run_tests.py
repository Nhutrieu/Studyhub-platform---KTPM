import os
import json
import shutil
import subprocess
from datetime import datetime, timezone

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

load_dotenv()

BASE_DIR = os.path.dirname(__file__)
PROJECT_DIR = os.path.dirname(BASE_DIR)

BASE_URL = os.getenv("BASE_URL", "http://localhost/HeThongChamSocCaKoi/backend")

JIRA_BASE_URL = os.getenv("JIRA_BASE_URL")
JIRA_EMAIL = os.getenv("JIRA_EMAIL")
JIRA_API_TOKEN = os.getenv("JIRA_API_TOKEN")
JIRA_PROJECT_KEY = os.getenv("JIRA_PROJECT_KEY", "KOI")
JIRA_ISSUE_TYPE = os.getenv("JIRA_ISSUE_TYPE", "Bug")

# Epic cha
JIRA_EPIC_KEY = os.getenv("JIRA_EPIC_KEY", "KOI-24")

COLLECTION = os.path.join(PROJECT_DIR, "postman", "koi_collection.json")

REPORT_DIR = os.path.join(PROJECT_DIR, "reports")
REPORT_JSON = os.path.join(REPORT_DIR, "report.json")
REPORT_MD = os.path.join(REPORT_DIR, "result.md")
REPORT_XLSX = os.path.join(REPORT_DIR, "report.xlsx")

os.makedirs(REPORT_DIR, exist_ok=True)


def run_newman():
    newman = shutil.which("newman")

    if not newman:
        npx = shutil.which("npx")
        if not npx:
            print("Không tìm thấy Newman.")
            print("Chạy: npm install -g newman")
            return False

    if newman:
        cmd = [newman]
    else:
        cmd = [npx, "newman"]

    cmd += [
        "run",
        COLLECTION,
        "--env-var",
        f"baseUrl={BASE_URL}",
        "--reporters",
        "cli,json",
        "--reporter-json-export",
        REPORT_JSON
    ]

    print("Running Newman...")
    result = subprocess.run(cmd)

    if result.returncode != 0:
        print("Newman có test FAIL. Vẫn tiếp tục tạo report và Jira bug...")

    return os.path.exists(REPORT_JSON)


def parse_report():
    with open(REPORT_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)

    results = []
    failures = []

    for ex in data.get("run", {}).get("executions", []):
        name = ex.get("item", {}).get("name", "Unknown Test")
        request = ex.get("request", {})
        method = request.get("method", "UNKNOWN")

        url_obj = request.get("url", {})
        url = url_obj.get("raw", "") if isinstance(url_obj, dict) else str(url_obj)

        response = ex.get("response", {})
        status = response.get("code", "NO RESPONSE")

        assertion_errors = []

        for assertion in ex.get("assertions", []):
            error = assertion.get("error")
            if error:
                assertion_errors.append(
                    f"{assertion.get('assertion')}: {error.get('message')}"
                )

        # Chỉ FAIL khi assertion fail
        result = "FAIL" if assertion_errors else "PASS"

        row = {
            "name": name,
            "method": method,
            "url": url,
            "status": status,
            "result": result,
            "message": "; ".join(assertion_errors)
        }

        results.append(row)

        if result == "FAIL":
            failures.append(row)

    return results, failures


def write_markdown(results, failures, created_issues):
    now = datetime.now(timezone.utc).isoformat()

    lines = [
        "# KOI Care API Testing Report",
        "",
        f"Generated: {now}",
        "",
        "## Test Summary",
        "",
        "| Test Case | Method | URL | Status | Result | Message |",
        "|---|---|---|---|---|---|"
    ]

    for r in results:
        lines.append(
            f"| {r['name']} | {r['method']} | {r['url']} | {r['status']} | {r['result']} | {r['message']} |"
        )

    lines.append("")
    lines.append("## Jira Issues")
    lines.append("")

    if created_issues:
        for issue in created_issues:
            lines.append(f"- {issue}")
    else:
        lines.append("No Jira issues created.")

    lines.append("")
    lines.append("## Conclusion")
    lines.append("")

    if failures:
        lines.append("Some API test cases failed. Jira bugs were created in the Epic.")
    else:
        lines.append("All API test cases passed successfully.")

    with open(REPORT_MD, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def write_excel(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "API Summary"

    blue_fill = PatternFill("solid", fgColor="00008B")
    white_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    headers = [
        "No",
        "Function Name",
        "Passed",
        "Failed",
        "Untested",
        "N",
        "A",
        "B",
        "Total Test Cases"
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border

    for index, r in enumerate(results, start=1):
        passed = 1 if r["result"] == "PASS" else 0
        failed = 1 if r["result"] == "FAIL" else 0

        ws.append([
            index,
            r["name"],
            passed,
            failed,
            0,
            0,
            0,
            0,
            1
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = center

    widths = {
        "A": 8,
        "B": 45,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 8,
        "G": 8,
        "H": 8,
        "I": 18
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    total_row = len(results) + 2

    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = Font(bold=True)

    ws[f"C{total_row}"] = f"=SUM(C2:C{len(results)+1})"
    ws[f"D{total_row}"] = f"=SUM(D2:D{len(results)+1})"
    ws[f"E{total_row}"] = f"=SUM(E2:E{len(results)+1})"
    ws[f"I{total_row}"] = f"=SUM(I2:I{len(results)+1})"

    for cell in ws[total_row]:
        cell.border = border
        cell.alignment = center

    wb.save(REPORT_XLSX)


def create_jira_issue(failure):
    if not all([JIRA_BASE_URL, JIRA_EMAIL, JIRA_API_TOKEN, JIRA_PROJECT_KEY]):
        print("Thiếu thông tin Jira trong .env.")
        return None

    url = f"{JIRA_BASE_URL.rstrip('/')}/rest/api/3/issue"

    description_text = (
        f"KOI Care API test failed.\n\n"
        f"Test Case: {failure['name']}\n"
        f"Method: {failure['method']}\n"
        f"Endpoint: {failure['url']}\n"
        f"Status Code: {failure['status']}\n"
        f"Error: {failure['message']}\n\n"
        f"Report files:\n"
        f"- reports/result.md\n"
        f"- reports/report.xlsx\n"
        f"- reports/report.json"
    )

    fields = {
        "project": {
            "key": JIRA_PROJECT_KEY
        },

       
        "parent": {
            "key": JIRA_EPIC_KEY
        },

        "summary": f"[API TEST FAILED] {failure['name']}",

        "description": {
            "type": "doc",
            "version": 1,
            "content": [
                {
                    "type": "paragraph",
                    "content": [
                        {
                            "type": "text",
                            "text": description_text
                        }
                    ]
                }
            ]
        },

        "issuetype": {
            "name": JIRA_ISSUE_TYPE
        }
    }

    payload = {
        "fields": fields
    }

    response = requests.post(
        url,
        json=payload,
        auth=(JIRA_EMAIL, JIRA_API_TOKEN),
        headers={
            "Accept": "application/json",
            "Content-Type": "application/json"
        }
    )

    if response.status_code in [200, 201]:
        issue_key = response.json().get("key")
        print(f"Created Jira bug: {issue_key}")
        print(f"Linked under epic: {JIRA_EPIC_KEY}")
        return issue_key

    print("Jira create failed:")
    print(response.status_code)
    print(response.text)

    return None


def main():
    if not run_newman():
        print("Không tạo được report.json")
        return

    results, failures = parse_report()
    created_issues = []

    if failures:
        print("Có API FAIL, đang tạo Jira Bug trong Epic...")
        for failure in failures:
            issue_key = create_jira_issue(failure)
            if issue_key:
                created_issues.append(issue_key)
    else:
        print("Tất cả API đều PASS.")

    write_markdown(results, failures, created_issues)
    write_excel(results)

    print("Done.")
    print(f"Markdown report: {REPORT_MD}")
    print(f"Excel report: {REPORT_XLSX}")
    print(f"JSON report: {REPORT_JSON}")


if __name__ == "__main__":
    main()